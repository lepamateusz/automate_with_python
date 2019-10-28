#! /usr/bin/python3

# multiplicationTable.py - takes a number form the command line and creates NXN muliplication table in an Exel spreadsheet
# usege schema: multiplicationTable.py N

import sys, openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

if len(sys.argv) != 2:
    print("Too many arguments!")
else:
    N = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active


for row in range(N):
    sheet['A'+str(row+2)].value = row + 1
    sheet['A' + str(row + 2)].font = Font(bold=True)
    for column in range(N):
        sheet[get_column_letter(column + 2) + str(1)].value = column + 1
        sheet[get_column_letter(column + 2) + str(1)].font = Font(bold=True)
        sheet[get_column_letter(row+2)+str(column+2)].value = (row+1)*(column+1)

wb.save('multiplicationTable.xlsx')




