#! /usr/bin/python3

# excelToCSVconverter.py - reads all the Excel files in the current working directory and outputs them as CSV files
# CSV format: <excel filename>_<sheet title>.csv

import openpyxl, os, csv

for excelFile in os.listdir('.'):
    # Skip non-xlsx files, load the workbook object.
    if not (excelFile.endswith('.xlsx') or excelFile.endswith('.xls')) :
        continue
    wb = openpyxl.load_workbook(excelFile)
    for sheetName in wb.get_sheet_names():
        # Loop through every sheet in the workbook.
        sheet = wb.get_sheet_by_name(sheetName)

        # Create the CSV filename from the Excel filename and sheet title.
        name = excelFile.split(".")[0]
        csvName = name + "_" + sheetName + '.csv'
        print(csvName)
        csvFile = open(csvName, 'w', newline='')
        # Create the csv.writer object for this CSV file.
        csvWriter = csv.writer(csvFile)
        # Loop through every row in the sheet.
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # append each cell to this list
            # Loop through each cell in the row.
            for colNum in range(1, sheet.max_column + 1):
                # Append each cell's data to rowData.
                rowData.append(colNum)

            # Write the rowData list to the CSV file.
            csvWriter.writerow(rowData)
        csvFile.close()